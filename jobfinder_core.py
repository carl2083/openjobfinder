import json
import os
import random
import re
from pathlib import Path
from zipfile import BadZipFile
from dataclasses import asdict
from dataclasses import dataclass
from dataclasses import fields
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.dimensions import RowDimension


DEFAULT_CONFIG_PATH = "config.json"
SKILL_PROFILE_PATH = "skill.md"
PROMPT_TEMPLATE_PATH = "prompt.md"
SKILL_PROFILE_MARKER = "<!-- GENERATED_SKILL_PROFILE_START -->"
RESUME_STYLE_OPTIONS = [
    "ATS-Focused Professional",
    "Concise Achievement-Led",
    "Project & Impact Focused",
]
DEFAULT_COVER_LETTER_STYLE = "Commercial Problem-Solver"
COVER_LETTER_STYLE_OPTIONS = [
    "Commercial Problem-Solver",
    "Evidence-Led Value Pitch",
    "Operational Calm & Reliability",
    "Strategic Career Transition",
]

# Cover letter visual DESIGN (PDF layout) — distinct from the writing-tone
# style above. Maps a selectable design name to its HTML template.
DEFAULT_COVER_LETTER_DESIGN = "Minimal Cover Letter"
COVER_LETTER_DESIGN_OPTIONS = [
    "Minimal Cover Letter",
    "Navy Gold",
]
COVER_LETTER_DESIGN_TEMPLATES = {
    "Minimal Cover Letter": "templates/cover-letter-template.html",
    "Navy Gold": "templates/navy-gold-template.html",
}


DEFAULT_PROMPT_TEMPLATE = """
You are evaluating job descriptions for me.
Please deeply analyse this job posting, the company and its industry. 
based on your analysis, identify the most important 9 - 12 skills that are most relevant to the role, 
and list those as 1-2 word long keywords sets that I can include in my resume. 

EXPERIENCE CONSTRAINT (CRITICAL):
- ONLY include job roles, employers, and position titles that I explicitly provide.
- DO NOT invent, rename, generalise, or group experience into new roles (e.g. “Residential Projects”, “Consulting Work”, “Various Clients”).
- If experience does not belong to a formal employer or job title, it must be labelled exactly as provided by me (e.g. “Business Owner – Café”, “Self-Directed Projects”).
- DO NOT create new time ranges or fill perceived employment gaps with inferred roles.


Then write me a resume IMPORTANT RULES (DO NOT IGNORE):
- DO NOT make up, exaggerate, or fabricate any experience, responsibilities, tools, systems, or results.
- ONLY use experience that can be reasonably inferred from my actual background.
- You may reframe, emphasize, or de-emphasize aspects of my real experience to best match the job description, but never introduce new experience.
- If a JD requirement is not directly met, use the closest transferable or adjacent experience instead and clearly reflect it truthfully.
- For EACH job experience, write 6–10 bullet points (never fewer than 6).
- Bullet points MUST be tailored to the job description and map directly to its responsibilities and requirements.
- You may reframe or emphasize different aspects of my experience to match the JD, but do NOT invent experience.
- Every bullet point MUST follow the X–Y–Z formula:
  - X = what was accomplished
  - Y = how it was accomplished (tools, methods, stakeholders)
  - Z = measurable outcome or business impact (%, $, time, scale). If exact metrics are unavailable, use reasonable qualitative impact.

ALIGNMENT REQUIREMENTS:
- Ensure ALL major responsibilities in the JD are reflected somewhere in the bullet points.
- If a JD requirement is missing from my background, show adjacent or transferable experience instead.
- Prioritize JD keywords, tools, systems, and methodologies in bullet points.

EXPERIENCE DEPTH:
- In the Professional Summary, explicitly state total years of relevant experience if applicable.
- Senior or long-term roles should have more bullet points than junior roles.

QUALITY BAR:
- No generic bullets.
- No duplicated ideas.
- No filler language.
- Bullets must sound like a strong Australian-market resume.

If fewer than 6 strong bullets exist for a role, derive additional bullets by:
- Breaking complex work into sub-achievements
- Highlighting stakeholder management, risk, compliance, reporting, or optimisation work

Resume must have the following format: 
    - Professional Summary: 
    - Skills: 
    - Experience: 
    - Education: 
    - Certifications: 
    - Languages: 
    - References: 
    (not strictly follow the format, just use it as a reference)
And write me a cover letter try answer the questions in JD.

COVER LETTER RULES:
- The cover letter should ideally fit on a single page.
- Keep it generally between 250 and 400 words.
- Structure it into 3 to 6 paragraphs.
- Focus on the strongest qualifications and fit for this role.
- Do not repeat the entire resume or restate every bullet point.
- Use specific, relevant evidence from my background that best supports this application.
- Keep it sharp, persuasive, and readable in an Australian job-market tone.

For bullet points use x,y,z formula: which is what you accomplished. and how you accomplished it. Then write me a resume and cover letter
Return STRICT JSON using the provided schema.
{
    "job_meta": {
        "job_title": "",
        "company": "",
        "location": "",
        "job_url": ""
    },
    "suitability": {
        "core_skill_match": 0,
        "tools_systems_overlap": 0,
        "industry_gatekeeping": 0,
        "seniority_fit": 0,
        "ats_keyword_match": 0,
        "location_logistics": 0
    },
    "interest": {
        "nature_of_work": 0,
        "structure_vs_chaos": 0,
        "learning_value": 0,
        "energy_drain": 0,
        "exit_value": 0
    },
    "notes": {
        "top_strengths": ["", ""],
        "main_risks": ["", ""],
        "resume_focus": ["", ""]
    },
    "other": {
        "resume_sections": {
            "name": "",
            "position": "",
            "address": "",
            "phone": "",
            "email": "",
            "professional_summary": "",
            "experience": [
                {
                    "company": "",
                    "location": "",
                    "role": "",
                    "date": "",
                    "bullets": ["", ""]
                }
            ],
            "education": [
                {
                    "institution": "",
                    "location": "",
                    "degree": ""
                }
            ],
            "skills": ["", ""]
        },
        "Cover Letter": ""
    }
}
Do not add commentary.
All scores must follow the defined scoring model.
"""


DEFAULT_SKILL_PROFILE_TEMPLATE = f"""# Copy This To ChatGPT

Copy the prompt below into ChatGPT together with your resume, work history, project history, or any rough notes about your background.
Ask ChatGPT to return a clean, truthful skill profile in markdown.

---

Please help me build a truthful master skill profile for job applications.

Rules:
- Only use experience, roles, projects, education, tools, and skills that I explicitly provide.
- Do not invent employers, achievements, systems, dates, or seniority.
- Rewrite my background into a clean markdown profile that is easy for another ChatGPT prompt to reuse later.
- Group similar experience clearly, but do not fabricate new jobs or fake responsibilities.
- Include:
  - Core Professional Background
  - Education
  - Business / Commercial Experience
  - Software / Technical Projects
  - Data / Analytics / Automation Skills
  - AI / Tooling
  - Technical Mindset & Strengths
  - Communication & Work Style
  - Explicit Exclusions
  - Positioning Guidance

Output requirements:
- Use markdown headings and bullet points.
- Keep it factual and reusable.
- Make it strong, but never exaggerated.
- This document will later be used as the source of truth for resume and cover-letter generation.

---

Paste the final generated profile below this marker:

{SKILL_PROFILE_MARKER}

"""


SHEET_HEADERS = [
    "Job ID",
    "Job Title",
    "Company",
    "Location",
    "Job URL",
    "Core Skill Match (0–30)",
    "Tools / Systems (0–20)",
    "Industry Gatekeeping (0–15)",
    "Seniority Fit (0–15)",
    "ATS Keyword Match (0–10)",
    "Location / Logistics (0–10)",
    "Suitability Total",
    "Nature of Work (0–30)",
    "Structure vs Chaos (0–20)",
    "Learning Value (0–20)",
    "Energy Drain (0–20) 0 is most draining, 20 is least draining",
    "Exit Value (0–10)",
    "Interest Total",
    "Final Score",
    "Recommendation (APPLY/MAYBE/SKIP)",
    "Confidence",
    "Top Strengths",
    "Main Risks",
    "Resume Focus",
    "Applied? (Y/N)",
    "Interview Outcome",
    "Notes",
    "Resume",
    "Cover Letter",
    "PDF Folder Link",
]

HIDDEN_EXCEL_COLUMNS = [chr(code) for code in range(ord("F"), ord("X") + 1)] + ["AB", "AC"]


@dataclass
class Config:
    output_excel: str = "job_results.xlsx"
    job_location: str = "All Gold Coast QLD"
    keyword: Optional[str] = "accountant"
    gpt_mode: str = "web_chatgpt"
    max_runs: Optional[int] = 20
    single_job_url: Optional[str] = None
    include_recommendations: bool = False
    include_new_to_you: bool = False
    exit_when_done: bool = False
    skip_title_contains: str = ""
    delay_between_jobs_min_sec: int = 10
    delay_between_jobs_max_sec: int = 20
    batch_size: int = 1
    apply_threshold: int = 140
    maybe_threshold: int = 100
    default_confidence: str = "MEDIUM"
    chrome_debug_port: int = 9222
    chrome_user_data_dir: str = ""
    chrome_path: str = ""
    seek_url: str = "https://www.seek.com.au/"
    chatgpt_url: str = "https://chat.openai.com/"
    chatgpt_chat_title: str = "Job application advice"
    enable_local_sync: bool = False
    local_sync_path: str = ""
    local_sync_pull_before_run: bool = True
    enable_pdf_export: bool = True
    pdf_css_path: str = ""
    pdf_output_dir: str = "pdf_output"
    pdf_template_path: str = "templates/navy-gold-template.html"
    user_address: str = ""
    user_phone: str = ""
    user_email: str = ""
    user_name: str = "carl chen"
    resume_style: str = "ATS-Focused Professional"
    cover_letter_style: str = DEFAULT_COVER_LETTER_STYLE
    cover_letter_design: str = DEFAULT_COVER_LETTER_DESIGN
    skill_profile_path: str = SKILL_PROFILE_PATH
    prompt_template_path: str = PROMPT_TEMPLATE_PATH
    profile_id: str = "default"
    profile_label: str = "Default"


@dataclass
class ProfileEntry:
    id: str
    label: str
    directory: str


@dataclass
class AppConfig:
    active_profile_id: str = "default"
    profiles: List[ProfileEntry] = None  # type: ignore[assignment]
    chrome_debug_port: int = 9222
    chrome_user_data_dir: str = ""
    chrome_path: str = ""
    initial_login_completed: bool = False

    def __post_init__(self) -> None:
        if self.profiles is None:
            self.profiles = [ProfileEntry(id="default", label="Default", directory="profiles/default")]


def _workspace_root(config_path: str) -> Path:
    return Path(config_path).resolve().parent


def _sanitize_profile_id(label: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", (label or "").strip().lower()).strip("-")
    return normalized or "profile"


def _dedupe_profile_id(existing_ids: List[str], preferred_id: str) -> str:
    if preferred_id not in existing_ids:
        return preferred_id
    counter = 2
    while f"{preferred_id}-{counter}" in existing_ids:
        counter += 1
    return f"{preferred_id}-{counter}"


def _profile_directory(profile_id: str) -> str:
    return str(Path("profiles") / profile_id)


def _default_output_excel(profile_id: str) -> str:
    return str(Path("profiles") / profile_id / f"job_results-{profile_id}.xlsx")


def _default_pdf_output_dir(profile_id: str) -> str:
    return str(Path("profiles") / profile_id / "pdf_output")


def _default_skill_profile_path(profile_id: str) -> str:
    return str(Path("profiles") / profile_id / "skill.md")


def _default_prompt_template_path(profile_id: str) -> str:
    return str(Path("profiles") / profile_id / "prompt.md")


def _default_profile_config(profile_id: str, label: str) -> Config:
    return Config(
        output_excel=_default_output_excel(profile_id),
        pdf_output_dir=_default_pdf_output_dir(profile_id),
        skill_profile_path=_default_skill_profile_path(profile_id),
        prompt_template_path=_default_prompt_template_path(profile_id),
        profile_id=profile_id,
        profile_label=label,
    )


def _profile_config_path(app_config: AppConfig, profile_id: str) -> Path:
    profile = next((p for p in app_config.profiles if p.id == profile_id), None)
    if profile is None:
        profile = app_config.profiles[0]
    return Path(profile.directory) / "profile.json"


def _app_config_to_dict(app_config: AppConfig) -> Dict[str, Any]:
    return {
        "active_profile_id": app_config.active_profile_id,
        "profiles": [asdict(profile) for profile in app_config.profiles],
        "chrome_debug_port": app_config.chrome_debug_port,
        "chrome_user_data_dir": app_config.chrome_user_data_dir,
        "chrome_path": app_config.chrome_path,
        "initial_login_completed": app_config.initial_login_completed,
    }


def _config_from_legacy_data(data: Dict[str, Any]) -> Config:
    allowed = {field.name for field in fields(Config)}
    sanitized = {key: value for key, value in data.items() if key in allowed}
    config = Config(**sanitized)
    if not getattr(config, "skill_profile_path", ""):
        config.skill_profile_path = SKILL_PROFILE_PATH
    if not getattr(config, "prompt_template_path", ""):
        config.prompt_template_path = PROMPT_TEMPLATE_PATH
    return config


def _is_app_config_shape(data: Dict[str, Any]) -> bool:
    return "profiles" in data and isinstance(data.get("profiles"), list)


def load_app_config(path: str = DEFAULT_CONFIG_PATH) -> AppConfig:
    root = _workspace_root(path)
    config_path = Path(path)
    if not config_path.exists():
        return AppConfig()
    with config_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not _is_app_config_shape(data):
        return migrate_legacy_config(path)

    profiles: List[ProfileEntry] = []
    for item in data.get("profiles", []):
        if not isinstance(item, dict):
            continue
        profile_id = str(item.get("id", "")).strip() or "default"
        label = str(item.get("label", "")).strip() or profile_id.title()
        directory = str(item.get("directory", "")).strip() or _profile_directory(profile_id)
        profiles.append(ProfileEntry(id=profile_id, label=label, directory=directory))
    if not profiles:
        profiles = [ProfileEntry(id="default", label="Default", directory=_profile_directory("default"))]

    app_config = AppConfig(
        active_profile_id=str(data.get("active_profile_id", profiles[0].id) or profiles[0].id),
        profiles=profiles,
        chrome_debug_port=int(data.get("chrome_debug_port", 9222) or 9222),
        chrome_user_data_dir=str(data.get("chrome_user_data_dir", "") or ""),
        chrome_path=str(data.get("chrome_path", "") or ""),
        initial_login_completed=bool(data.get("initial_login_completed", False)),
    )

    if app_config.active_profile_id not in {profile.id for profile in app_config.profiles}:
        app_config.active_profile_id = app_config.profiles[0].id

    for profile in app_config.profiles:
        profile_dir = root / profile.directory
        profile_dir.mkdir(parents=True, exist_ok=True)
    return app_config


def save_app_config(path: str, app_config: AppConfig) -> None:
    root = _workspace_root(path)
    for profile in app_config.profiles:
        (root / profile.directory).mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(_app_config_to_dict(app_config), f, indent=2)


def load_profile_config(
    app_config: AppConfig,
    profile_id: Optional[str] = None,
    config_path: str = DEFAULT_CONFIG_PATH,
) -> Config:
    root = _workspace_root(config_path)
    target_profile_id = profile_id or app_config.active_profile_id
    profile = next((p for p in app_config.profiles if p.id == target_profile_id), None)
    if profile is None:
        profile = app_config.profiles[0]
        target_profile_id = profile.id

    path_obj = root / _profile_config_path(app_config, target_profile_id)
    if path_obj.exists():
        with path_obj.open("r", encoding="utf-8") as f:
            data = json.load(f)
        config = _config_from_legacy_data(data)
    else:
        config = _default_profile_config(profile.id, profile.label)

    config.profile_id = profile.id
    config.profile_label = profile.label
    config.chrome_debug_port = app_config.chrome_debug_port
    config.chrome_user_data_dir = app_config.chrome_user_data_dir
    config.chrome_path = app_config.chrome_path
    if not config.skill_profile_path:
        config.skill_profile_path = _default_skill_profile_path(profile.id)
    if not config.prompt_template_path:
        config.prompt_template_path = _default_prompt_template_path(profile.id)
    if not config.output_excel:
        config.output_excel = _default_output_excel(profile.id)
    if not config.pdf_output_dir:
        config.pdf_output_dir = _default_pdf_output_dir(profile.id)
    return config


def save_profile_config(
    app_config: AppConfig,
    config: Config,
    profile_id: Optional[str] = None,
    config_path: str = DEFAULT_CONFIG_PATH,
) -> None:
    root = _workspace_root(config_path)
    target_profile_id = profile_id or config.profile_id or app_config.active_profile_id
    profile = next((p for p in app_config.profiles if p.id == target_profile_id), None)
    if profile is None:
        raise ValueError(f"Unknown profile id: {target_profile_id}")

    config.profile_id = profile.id
    config.profile_label = profile.label
    profile_dir = root / profile.directory
    profile_dir.mkdir(parents=True, exist_ok=True)

    config.chrome_debug_port = app_config.chrome_debug_port
    config.chrome_user_data_dir = app_config.chrome_user_data_dir
    config.chrome_path = app_config.chrome_path
    if not config.skill_profile_path:
        config.skill_profile_path = _default_skill_profile_path(profile.id)
    if not config.prompt_template_path:
        config.prompt_template_path = _default_prompt_template_path(profile.id)
    if not config.output_excel:
        config.output_excel = _default_output_excel(profile.id)
    if not config.pdf_output_dir:
        config.pdf_output_dir = _default_pdf_output_dir(profile.id)

    profile_path = root / _profile_config_path(app_config, profile.id)
    with profile_path.open("w", encoding="utf-8") as f:
        json.dump(asdict(config), f, indent=2)


def migrate_legacy_config(path: str = DEFAULT_CONFIG_PATH) -> AppConfig:
    root = _workspace_root(path)
    config_path = Path(path)
    legacy_data: Dict[str, Any] = {}
    if config_path.exists():
        with config_path.open("r", encoding="utf-8") as f:
            legacy_data = json.load(f)

    legacy_config = _config_from_legacy_data(legacy_data) if legacy_data else Config()
    profile_id = "default"
    profile_label = "Default"
    profile_dir = root / _profile_directory(profile_id)
    profile_dir.mkdir(parents=True, exist_ok=True)

    if not legacy_config.skill_profile_path or legacy_config.skill_profile_path == SKILL_PROFILE_PATH:
        source_skill = root / SKILL_PROFILE_PATH
        if source_skill.exists():
            target = profile_dir / "skill.md"
            if not target.exists():
                target.write_text(source_skill.read_text(encoding="utf-8"), encoding="utf-8")
        legacy_config.skill_profile_path = str(Path(_profile_directory(profile_id)) / "skill.md")
    if not legacy_config.prompt_template_path or legacy_config.prompt_template_path == PROMPT_TEMPLATE_PATH:
        source_prompt = root / PROMPT_TEMPLATE_PATH
        if source_prompt.exists():
            target = profile_dir / "prompt.md"
            if not target.exists():
                target.write_text(source_prompt.read_text(encoding="utf-8"), encoding="utf-8")
        legacy_config.prompt_template_path = str(Path(_profile_directory(profile_id)) / "prompt.md")

    legacy_config.profile_id = profile_id
    legacy_config.profile_label = profile_label
    app_config = AppConfig(
        active_profile_id=profile_id,
        profiles=[ProfileEntry(id=profile_id, label=profile_label, directory=_profile_directory(profile_id))],
        chrome_debug_port=legacy_config.chrome_debug_port,
        chrome_user_data_dir=legacy_config.chrome_user_data_dir,
        chrome_path=legacy_config.chrome_path,
        initial_login_completed=False,
    )
    save_profile_config(app_config, legacy_config, profile_id=profile_id, config_path=path)
    save_app_config(path, app_config)
    return app_config


def create_profile(
    app_config: AppConfig,
    label: str,
    config_path: str = DEFAULT_CONFIG_PATH,
) -> ProfileEntry:
    existing_ids = [profile.id for profile in app_config.profiles]
    profile_id = _dedupe_profile_id(existing_ids, _sanitize_profile_id(label))
    profile = ProfileEntry(id=profile_id, label=label.strip() or profile_id.title(), directory=_profile_directory(profile_id))
    app_config.profiles.append(profile)
    config = _default_profile_config(profile.id, profile.label)
    save_app_config(config_path, app_config)
    save_profile_config(app_config, config, profile_id=profile.id, config_path=config_path)
    return profile


def rename_profile(
    app_config: AppConfig,
    profile_id: str,
    new_label: str,
    config_path: str = DEFAULT_CONFIG_PATH,
) -> None:
    profile = next((p for p in app_config.profiles if p.id == profile_id), None)
    if profile is None:
        raise ValueError(f"Unknown profile id: {profile_id}")
    profile.label = new_label.strip() or profile.label
    config = load_profile_config(app_config, profile_id=profile_id, config_path=config_path)
    config.profile_label = profile.label
    save_profile_config(app_config, config, profile_id=profile_id, config_path=config_path)
    save_app_config(config_path, app_config)


def delete_profile(
    app_config: AppConfig,
    profile_id: str,
    config_path: str = DEFAULT_CONFIG_PATH,
) -> None:
    if len(app_config.profiles) <= 1:
        raise ValueError("Cannot delete the last remaining profile.")
    profile = next((p for p in app_config.profiles if p.id == profile_id), None)
    if profile is None:
        raise ValueError(f"Unknown profile id: {profile_id}")
    root = _workspace_root(config_path)
    profile_dir = root / profile.directory
    if profile_dir.exists():
        for child in sorted(profile_dir.rglob("*"), reverse=True):
            if child.is_file():
                child.unlink()
            elif child.is_dir():
                child.rmdir()
        profile_dir.rmdir()
    app_config.profiles = [item for item in app_config.profiles if item.id != profile_id]
    if app_config.active_profile_id == profile_id:
        app_config.active_profile_id = app_config.profiles[0].id
    save_app_config(config_path, app_config)


def load_config(path: str) -> Config:
    app_config = load_app_config(path)
    return load_profile_config(app_config, config_path=path)


def save_config(path: str, config: Config) -> None:
    app_config = load_app_config(path)
    app_config.active_profile_id = config.profile_id or app_config.active_profile_id
    app_config.chrome_debug_port = config.chrome_debug_port
    app_config.chrome_user_data_dir = config.chrome_user_data_dir
    app_config.chrome_path = config.chrome_path
    save_app_config(path, app_config)
    save_profile_config(app_config, config, profile_id=app_config.active_profile_id, config_path=path)


def load_job_ids_from_excel(path: str) -> List[str]:
    if not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path, read_only=True)
    except (BadZipFile, InvalidFileException):
        # 文件存在但不是合法的 xlsx，避免直接崩溃
        return []
    ws = wb.active
    ids: List[str] = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        value = row[0]
        if value:
            ids.append(str(value).strip())
    return ids


def extract_job_id(job_url: str) -> Optional[str]:
    if not job_url:
        return None
    match = re.search(r"/job/(\d+)", job_url)
    if match:
        return match.group(1)
    # Seek 等站点可能用 /job/12345678，兜底：取 URL 中 6 位以上数字
    fallback = re.search(r"(\d{6,})", job_url)
    return fallback.group(1) if fallback else None


def compute_totals(payload: Dict[str, Any]) -> Dict[str, int]:
    suitability = payload.get("suitability") or {}
    interest = payload.get("interest") or {}
    suitability_total = (
        suitability.get("core_skill_match", 0)
        + suitability.get("tools_systems_overlap", 0)
        + suitability.get("industry_gatekeeping", 0)
        + suitability.get("seniority_fit", 0)
        + suitability.get("ats_keyword_match", 0)
        + suitability.get("location_logistics", 0)
    )
    interest_total = (
        interest.get("nature_of_work", 0)
        + interest.get("structure_vs_chaos", 0)
        + interest.get("learning_value", 0)
        + interest.get("energy_drain", 0)
        + interest.get("exit_value", 0)
    )
    final_score = suitability_total + interest_total
    return {
        "suitability_total": suitability_total,
        "interest_total": interest_total,
        "final_score": final_score,
    }


def recommendation_for(score: int, config: Config) -> str:
    if score >= config.apply_threshold:
        return "APPLY"
    if score >= config.maybe_threshold:
        return "MAYBE"
    return "SKIP"


def format_list(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if v)
    return str(value) if value is not None else ""


def ensure_prompt_template_file(path: str = PROMPT_TEMPLATE_PATH) -> str:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    if not os.path.exists(path):
        Path(path).write_text(DEFAULT_PROMPT_TEMPLATE.strip() + "\n", encoding="utf-8")
    return path


def ensure_skill_profile_file(path: str = SKILL_PROFILE_PATH) -> str:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    if not os.path.exists(path):
        Path(path).write_text(DEFAULT_SKILL_PROFILE_TEMPLATE, encoding="utf-8")
    return path


def load_skill_profile(path: str = SKILL_PROFILE_PATH) -> str:
    if not os.path.exists(path):
        return ""
    try:
        text = Path(path).read_text(encoding="utf-8").strip()
    except Exception:
        return ""
    if SKILL_PROFILE_MARKER in text:
        text = text.split(SKILL_PROFILE_MARKER, 1)[1].strip()
    return text


def load_prompt_template(path: str = PROMPT_TEMPLATE_PATH) -> str:
    if not os.path.exists(path):
        return DEFAULT_PROMPT_TEMPLATE.strip()
    try:
        text = Path(path).read_text(encoding="utf-8").strip()
    except Exception:
        return DEFAULT_PROMPT_TEMPLATE.strip()
    return text or DEFAULT_PROMPT_TEMPLATE.strip()


def _writing_style_block(config: Config) -> str:
    return """
WRITING STYLE PREFERENCES:
- Resume style: {resume_style}
- Cover letter style: {cover_letter_style}

Use the selected styles as tone and structure guidance while keeping everything truthful and tailored to the JD.
""".strip()


def _candidate_identity_block(config: Config) -> str:
    return """
CURRENT CANDIDATE SETTINGS:
- Profile label: {profile_label}
- Full name: {user_name}
- Address: {user_address}
- Phone: {user_phone}
- Email: {user_email}

CRITICAL IDENTITY RULES:
- Treat the candidate settings above as the authoritative identity for this run.
- Do not reuse any name, address, phone, or email from earlier chats or previous candidates.
- In `other.resume_sections`, set `name`, `address`, `phone`, and `email` from these settings.
- In the cover letter signature and contact block, use these same settings.
- If a setting is blank, leave that field blank instead of inventing a value.
""".strip()


def _truth_guard_block() -> str:
    return """
TRUTH AND FORMAT RULES:
- Do not infer or exaggerate years of experience.
- Do not write claims like `7+ years`, `5+ years`, or similar unless the candidate source material explicitly supports that exact duration.
- If exact years are unclear, describe the background qualitatively instead of inventing a number.
- Do not invent seniority, certifications, employers, projects, achievements, tools, or job titles.
- Capitalize personal names and job titles normally in the final output.

MANDATORY OUTPUT (DO NOT SKIP - APPLIES REGARDLESS OF SCORES):
- ALWAYS produce a complete, tailored resume AND a complete cover letter for EVERY job, even when the suitability or interest scores are low or you would otherwise recommend skipping the role.
- The suitability/interest scores and any skip recommendation are for my reference ONLY. They must NEVER cause you to leave resume or cover letter content empty.
- `other.resume_sections` must be fully populated every time: `professional_summary`, `experience` (each role with 6-10 X-Y-Z bullets), `education`, and `skills`. Never return empty strings or empty arrays for these fields.
- `other."Cover Letter"` must ALWAYS contain a full 250-400 word cover letter body. Never return it empty.
- If the role is a weak fit, still write the strongest possible TRUTHFUL resume and cover letter using transferable or adjacent experience - never fabricate to compensate.

COVER LETTER OUTPUT RULES:
- Return only the cover letter body content.
- You may include the salutation line such as `Dear Hiring Manager,` or a named addressee if the job ad clearly provides one.
- Do not include the candidate contact block at the top.
- Do not repeat the candidate name, address, phone, or email inside the body unless the letter naturally requires it.
- Do not include a closing signature block like `Kind regards` plus the candidate name.
""".strip()


def build_prompt(job_description: str, config: Config) -> str:
    prompt_template = load_prompt_template(getattr(config, "prompt_template_path", PROMPT_TEMPLATE_PATH))
    skill_profile = load_skill_profile(getattr(config, "skill_profile_path", SKILL_PROFILE_PATH))
    skill_block = ""
    if skill_profile:
        skill_block = (
            "SKILL PROFILE (source of truth; do not fabricate beyond this):\n"
            f"{skill_profile}\n\n"
        )
    style_block = _writing_style_block(config).format(
        resume_style=getattr(config, "resume_style", "ATS-Focused Professional"),
        cover_letter_style=getattr(
            config, "cover_letter_style", DEFAULT_COVER_LETTER_STYLE
        ),
    )
    candidate_block = _candidate_identity_block(config).format(
        profile_label=getattr(config, "profile_label", ""),
        user_name=getattr(config, "user_name", "") or "",
        user_address=getattr(config, "user_address", "") or "",
        user_phone=getattr(config, "user_phone", "") or "",
        user_email=getattr(config, "user_email", "") or "",
    )
    truth_block = _truth_guard_block()
    sections = [prompt_template.strip()]
    sections.append(style_block)
    sections.append(candidate_block)
    sections.append(truth_block)
    if skill_block:
        sections.append(skill_block.rstrip())
    sections.append(f"JOB DESCRIPTION:\n{job_description}")
    return "\n\n".join(section for section in sections if section)


def extract_json_from_text(text: str) -> Dict[str, Any]:
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
    if fenced:
        return json.loads(fenced.group(1))

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("No JSON object found in ChatGPT response.")
    return json.loads(text[start : end + 1])


def build_row(payload: Dict[str, Any], config: Config) -> List[Any]:
    job_meta = payload.get("job_meta") or {}
    suitability = payload.get("suitability") or {}
    interest = payload.get("interest") or {}
    notes = payload.get("notes", {})
    other = payload.get("other", {})

    job_id = extract_job_id(job_meta.get("job_url", "")) or job_meta.get("job_id")
    if not job_id:
        job_url = job_meta.get("job_url", "")
        job_id = (job_url.rstrip("/").split("/")[-1] or "unknown") if job_url else "unknown"

    totals = compute_totals(payload)
    recommendation = recommendation_for(totals["final_score"], config)

    pdf_folder_path = os.path.abspath(os.path.join(config.pdf_output_dir, str(job_id)))
    pdf_folder_link = f"file:///{pdf_folder_path.replace(os.sep, '/')}"

    return [
        job_id,
        job_meta.get("job_title", ""),
        job_meta.get("company", ""),
        job_meta.get("location", ""),
        job_meta.get("job_url", ""),
        suitability.get("core_skill_match", 0),
        suitability.get("tools_systems_overlap", 0),
        suitability.get("industry_gatekeeping", 0),
        suitability.get("seniority_fit", 0),
        suitability.get("ats_keyword_match", 0),
        suitability.get("location_logistics", 0),
        totals["suitability_total"],
        interest.get("nature_of_work", 0),
        interest.get("structure_vs_chaos", 0),
        interest.get("learning_value", 0),
        interest.get("energy_drain", 0),
        interest.get("exit_value", 0),
        totals["interest_total"],
        totals["final_score"],
        recommendation,
        config.default_confidence,
        format_list(notes.get("top_strengths")),
        format_list(notes.get("main_risks")),
        format_list(notes.get("resume_focus")),
        "",
        "",
        format_list(notes.get("notes")),
        format_list(other.get("Resume", notes.get("Resume"))),
        format_list(other.get("Cover Letter", notes.get("Cover Letter"))),
        pdf_folder_link,
    ]


def ensure_workbook(path: str) -> Workbook:
    Path(path).expanduser().parent.mkdir(parents=True, exist_ok=True)
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            apply_excel_view(wb.active)
            return wb
        except (BadZipFile, InvalidFileException):
            # 备份损坏文件，避免覆盖并保留现场
            ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
            backup_path = f"{path}.corrupt-{ts}.xlsx"
            os.replace(path, backup_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(SHEET_HEADERS)
    apply_excel_view(ws)
    wb.save(path)
    return wb


def apply_excel_view(ws) -> None:
    for col in HIDDEN_EXCEL_COLUMNS:
        ws.column_dimensions[col].hidden = True


def _parse_skip_title_strings(s: str) -> List[str]:
    if not s or not s.strip():
        return []
    # 支持英文逗号 , 和中文逗号 ，
    parts = re.split(r"[,，]", s)
    return [p.strip().lower() for p in parts if p.strip()]


def _job_title_matches_skip(job_title: str, skip_strings: List[str]) -> Optional[str]:
    if not skip_strings:
        return None
    title_lower = (job_title or "").lower()
    title_words = re.findall(r"[a-z0-9]+", title_lower)
    for k in skip_strings:
        if not k:
            continue
        if k in title_lower:
            return k
        for w in title_words:
            if k in w or w in k:
                return k
    return None


def should_skip_job_by_title(job_title: str, config: Config) -> Optional[str]:
    skip_strings = _parse_skip_title_strings(getattr(config, "skip_title_contains", "") or "")
    return _job_title_matches_skip(job_title, skip_strings)


def append_skipped_job_to_excel(
    config: Config, job_meta: Dict[str, Any], skip_reason: str
) -> str:
    job_id = extract_job_id(job_meta.get("job_url", "")) or job_meta.get("job_id")
    if not job_id:
        job_url = job_meta.get("job_url", "")
        job_id = (job_url.rstrip("/").split("/")[-1] or "unknown") if job_url else "unknown"
    payload = {
        "job_meta": {**job_meta, "job_id": job_id},
        "suitability": {},
        "interest": {},
        "notes": {"notes": [f"跳过: {skip_reason}"]},
        "other": {},
    }
    for k in ["core_skill_match", "tools_systems_overlap", "industry_gatekeeping",
              "seniority_fit", "ats_keyword_match", "location_logistics"]:
        payload["suitability"][k] = 0
    for k in ["nature_of_work", "structure_vs_chaos", "learning_value",
              "energy_drain", "exit_value"]:
        payload["interest"][k] = 0
    row = build_row(payload, config)
    row[26] = f"跳过: 标题包含 {skip_reason}"
    wb = ensure_workbook(config.output_excel)
    ws = wb.active
    ws.append(row)
    apply_excel_view(ws)
    row_num = ws.max_row
    dim = RowDimension(ws, index=row_num, hidden=True)
    ws.row_dimensions[row_num] = dim
    url_cell = ws.cell(row=row_num, column=5)
    if url_cell.value:
        url_cell.hyperlink = url_cell.value
        url_cell.style = "Hyperlink"
    wb.save(config.output_excel)
    return job_id


def append_row_to_excel(config: Config, payload: Dict[str, Any]) -> str:
    wb = ensure_workbook(config.output_excel)
    ws = wb.active
    row = build_row(payload, config)
    ws.append(row)
    apply_excel_view(ws)
    # 将 Job URL 设置为可点击的超链接
    url_cell = ws.cell(row=ws.max_row, column=5)
    if url_cell.value:
        url_cell.hyperlink = url_cell.value
        url_cell.style = "Hyperlink"
    # PDF 文件夹链接 (AD 列)
    pdf_link_cell = ws.cell(row=ws.max_row, column=30)
    if pdf_link_cell.value:
        pdf_link_cell.hyperlink = pdf_link_cell.value
        pdf_link_cell.style = "Hyperlink"
    wb.save(config.output_excel)
    return str(row[0])

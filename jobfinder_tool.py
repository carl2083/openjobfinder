import argparse
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: pip install openpyxl"
    ) from exc


DEFAULT_CONFIG_PATH = "config.json"
DEFAULT_OUTPUT = "job_results.xlsx"
DEFAULT_JOB_IDS = "job_ids.txt"


PROMPT_TEMPLATE = """You are evaluating job descriptions for me.
Please deeply analyse this job posting, the company and its industry. 
based on your analysis, identify the most important 9 - 12 skills that are most relevant to the role, 
and list those as 1-2 word long keywords sets that I can include in my resume. 

Then write me a resume IMPORTANT RULES (DO NOT IGNORE):
- For EACH job experience, write 6–10 bullet points (never fewer than 6).
- Bullet points MUST be tailored to the job description and map directly to its responsibilities and requirements.
- You may reframe or emphasize different aspects of my experience to match the JD, but do NOT invent experience.
- Every bullet point MUST follow the X–Y–Z formula:
  - X = what was accomplished
  - Y = how it was accomplished (tools, methods, stakeholders)
  - Z = measurable outcome or business impact (%, $, time, scale). If exact metrics are unavailable, use reasonable qualitative impact.

ALIGNMENT REQUIREMENTS:
- Ensure ALL major responsibilities in the JD are reflected somewhere in the bullet points.
- If a JD requirement is missing from my background, show adjacent or transferable experience instead.
- Prioritize JD keywords, tools, systems, and methodologies in bullet points.

EXPERIENCE DEPTH:
- In the Professional Summary, explicitly state total years of relevant experience if applicable.
- Senior or long-term roles should have more bullet points than junior roles.

QUALITY BAR:
- No generic bullets.
- No duplicated ideas.
- No filler language.
- Bullets must sound like a strong Australian-market resume.

If fewer than 6 strong bullets exist for a role, derive additional bullets by:
- Breaking complex work into sub-achievements
- Highlighting stakeholder management, risk, compliance, reporting, or optimisation work

Resume must have the following format: 
    - Professional Summary: 
    - Skills: 
    - Experience: 
    - Education: 
    - Certifications: 
    - Languages: 
    - References: 
    (not strictly follow the format, just use it as a reference)
And write me a cover letter try answer the questions in JD.
For bullet points use x,y,z formula: which is what you accomplished. and how you accomplished it. Then write me a resume and cover letter
Return STRICT JSON using the provided schema.
{
    "job_meta": {
        "job_title": "",
        "company": "",
        "location": "",
        "job_url": ""
    },
    "suitability": {
        "core_skill_match": 0,
        "tools_systems_overlap": 0,
        "industry_gatekeeping": 0,
        "seniority_fit": 0,
        "ats_keyword_match": 0,
        "location_logistics": 0
    },
    "interest": {
        "nature_of_work": 0,
        "structure_vs_chaos": 0,
        "learning_value": 0,
        "energy_drain": 0,
        "exit_value": 0
    },
    "notes": {
        "top_strengths": ["", ""],
        "main_risks": ["", ""],
        "resume_focus": ["", ""]
    },
    "other": {
        "resume_sections": {
            "name": "",
            "position": "",
            "address": "",
            "phone": "",
            "email": "",
            "professional_summary": "",
            "experience": [
                {
                    "company": "",
                    "location": "",
                    "role": "",
                    "date": "",
                    "bullets": ["", ""]
                }
            ],
            "education": [
                {
                    "institution": "",
                    "location": "",
                    "degree": ""
                }
            ],
            "skills": ["", ""]
        },
        "Cover Letter": ""
    }
}
Do not add commentary.
All scores must follow the defined scoring model.
"""


EXCEL_HEADERS = [
    "Job ID",
    "Job Title",
    "Company",
    "Location",
    "Job URL",
    "Core Skill Match (0–30)",
    "Tools / Systems (0–20)",
    "Industry Gatekeeping (0–15)",
    "Seniority Fit (0–15)",
    "ATS Keyword Match (0–10)",
    "Location / Logistics (0–10)",
    "Suitability Total",
    "Nature of Work (0–30)",
    "Structure vs Chaos (0–20)",
    "Learning Value (0–20)",
    "Energy Drain (0–20)",
    "Exit Value (0–10)",
    "Interest Total",
    "Final Score",
    "Recommendation (APPLY/MAYBE/SKIP)",
    "Confidence",
    "Top Strengths",
    "Main Risks",
    "Resume Focus",
    "Applied? (Y/N)",
    "Interview Outcome",
    "Notes",
    "Resume",
    "Cover Letter",
    "PDF Folder Link",
]


@dataclass
class Config:
    output_excel: str = DEFAULT_OUTPUT
    job_location: str = "All Gold Coast QLD"
    keyword: Optional[str] = None
    job_ids_file: str = DEFAULT_JOB_IDS
    gpt_mode: str = "web_manual"
    max_runs: Optional[int] = 5
    apply_threshold: int = 140
    maybe_threshold: int = 100
    default_confidence: str = "MEDIUM"


def load_config(path: str) -> Config:
    if not os.path.exists(path):
        return Config()
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return Config(**data)


def save_config(path: str, config: Config) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(config.__dict__, f, indent=2)


def load_job_ids(path: str) -> List[str]:
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def append_job_id(path: str, job_id: str) -> None:
    with open(path, "a", encoding="utf-8") as f:
        f.write(f"{job_id}\n")


def ensure_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        wb = load_workbook(path)
        return wb
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(EXCEL_HEADERS)
    wb.save(path)
    return wb


def extract_job_id(job_url: str) -> Optional[str]:
    if not job_url:
        return None
    match = re.search(r"/job/(\d+)", job_url)
    return match.group(1) if match else None


def compute_totals(payload: Dict[str, Any]) -> Dict[str, int]:
    suitability = payload["suitability"]
    interest = payload["interest"]
    suitability_total = (
        suitability["core_skill_match"]
        + suitability["tools_systems_overlap"]
        + suitability["industry_gatekeeping"]
        + suitability["seniority_fit"]
        + suitability["ats_keyword_match"]
        + suitability["location_logistics"]
    )
    interest_total = (
        interest["nature_of_work"]
        + interest["structure_vs_chaos"]
        + interest["learning_value"]
        + interest["energy_drain"]
        + interest["exit_value"]
    )
    final_score = suitability_total + interest_total
    return {
        "suitability_total": suitability_total,
        "interest_total": interest_total,
        "final_score": final_score,
    }


def recommendation_for(score: int, config: Config) -> str:
    if score >= config.apply_threshold:
        return "APPLY"
    if score >= config.maybe_threshold:
        return "MAYBE"
    return "SKIP"


def prompt_for_json() -> Dict[str, Any]:
    print("\n=== Copy this prompt to ChatGPT ===\n")
    print(PROMPT_TEMPLATE)
    print("\n=== Paste the JSON response below ===")
    print("Paste JSON, then enter a blank line to finish.\n")

    lines: List[str] = []
    while True:
        line = input()
        if not line.strip():
            break
        lines.append(line)

    raw = "\n".join(lines)
    return json.loads(raw)


def format_list(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if v)
    return str(value) if value is not None else ""


def append_row(workbook_path: str, payload: Dict[str, Any], config: Config) -> str:
    wb = ensure_workbook(workbook_path)
    ws = wb.active

    job_meta = payload["job_meta"]
    suitability = payload["suitability"]
    interest = payload["interest"]
    notes = payload.get("notes", {})
    other = payload.get("other", {})

    job_id = extract_job_id(job_meta.get("job_url", "")) or job_meta.get("job_id")
    if not job_id:
        job_id = f"manual-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    totals = compute_totals(payload)
    recommendation = recommendation_for(totals["final_score"], config)

    pdf_folder_path = os.path.abspath(os.path.join(config.pdf_output_dir, str(job_id)))
    pdf_folder_link = f"file:///{pdf_folder_path.replace(os.sep, '/')}"

    ws.append(
        [
            job_id,
            job_meta.get("job_title", ""),
            job_meta.get("company", ""),
            job_meta.get("location", ""),
            job_meta.get("job_url", ""),
            suitability.get("core_skill_match", 0),
            suitability.get("tools_systems_overlap", 0),
            suitability.get("industry_gatekeeping", 0),
            suitability.get("seniority_fit", 0),
            suitability.get("ats_keyword_match", 0),
            suitability.get("location_logistics", 0),
            totals["suitability_total"],
            interest.get("nature_of_work", 0),
            interest.get("structure_vs_chaos", 0),
            interest.get("learning_value", 0),
            interest.get("energy_drain", 0),
            interest.get("exit_value", 0),
            totals["interest_total"],
            totals["final_score"],
            recommendation,
            config.default_confidence,
            format_list(notes.get("top_strengths")),
            format_list(notes.get("main_risks")),
            format_list(notes.get("resume_focus")),
            "",
            "",
            format_list(notes.get("notes")),
            format_list(other.get("Resume", notes.get("Resume"))),
            format_list(other.get("Cover Letter", notes.get("Cover Letter"))),
            pdf_folder_link,
        ]
    )

    # 将 Job URL 设置为可点击的超链接
    url_cell = ws.cell(row=ws.max_row, column=5)
    if url_cell.value:
        url_cell.hyperlink = url_cell.value
        url_cell.style = "Hyperlink"
    # PDF 文件夹链接 (AD 列)
    pdf_link_cell = ws.cell(row=ws.max_row, column=30)
    if pdf_link_cell.value:
        pdf_link_cell.hyperlink = pdf_link_cell.value
        pdf_link_cell.style = "Hyperlink"

    wb.save(workbook_path)
    return job_id


def run_manual(config: Config) -> None:
    job_ids = set(load_job_ids(config.job_ids_file))
    runs_left = config.max_runs

    while runs_left is None or runs_left > 0:
        payload = prompt_for_json()
        job_meta = payload.get("job_meta", {})
        job_id = extract_job_id(job_meta.get("job_url", "")) or job_meta.get("job_id")

        if job_id and job_id in job_ids:
            print(f"Skipping already processed job: {job_id}")
        else:
            saved_id = append_row(config.output_excel, payload, config)
            append_job_id(config.job_ids_file, saved_id)
            job_ids.add(saved_id)
            print(f"Saved job: {saved_id}")

        if runs_left is not None:
            runs_left -= 1


def run_selenium(_: Config) -> None:
    raise NotImplementedError(
        "Selenium mode is not implemented yet. Use web_manual mode for now."
    )


def cmd_init(args: argparse.Namespace) -> None:
    config = Config()
    save_config(args.config, config)
    ensure_workbook(config.output_excel)
    if not os.path.exists(config.job_ids_file):
        open(config.job_ids_file, "a", encoding="utf-8").close()
    print(f"Created {args.config} and {config.output_excel}")


def cmd_run(args: argparse.Namespace) -> None:
    config = load_config(args.config)
    if config.gpt_mode == "web_manual":
        run_manual(config)
        return
    if config.gpt_mode == "selenium":
        run_selenium(config)
        return
    raise SystemExit(f"Unsupported gpt_mode: {config.gpt_mode}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="JobFinder tool runner")
    parser.add_argument(
        "--config", default=DEFAULT_CONFIG_PATH, help="Path to config.json"
    )
    sub = parser.add_subparsers(dest="command", required=True)

    init_cmd = sub.add_parser("init", help="Create config and workbook")
    init_cmd.set_defaults(func=cmd_init)

    run_cmd = sub.add_parser("run", help="Run in configured mode")
    run_cmd.set_defaults(func=cmd_run)

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
